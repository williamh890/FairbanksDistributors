from openpyxl import Workbook
from datetime import datetime, date, timedelta
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz
import math

# set fonts for sheet
small_font = Font(b=True, name='Arial', size=14)
medium_font = Font(b=True, name='Arial', size=18)
large_font = Font(b=True, name='Arial', size=24)
note_font = Font(name='Arial', size=22)
category_font = Font(b=True, name='Arial', size=28)
column_font_size = {0: 18, 1:24, 2:14, 3:22, 4:20}
column_fonts = {}
for column_num in column_font_size:
    column_fonts[column_num] = Font(b=True, name='Arial', size=column_font_size[column_num])

thin = Side(border_style="thin", color="000000")

box_border = Border(top=thin, left=thin, right=thin, bottom=thin)
side_border = Border(left=thin, right=thin)
side_with_bot_border = Border(left=thin, right=thin, bottom=thin)
center = Alignment(horizontal='center')

category_color = PatternFill("solid", fgColor="dddddd")
top_color = PatternFill("solid", fgColor="999999")

number_of_rows_per_page = 39


def write_xlsx(dest_dir, order_info):

    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    number_of_rows = len(order_info['items']) + len(set([category["type"] for category in order_info["items"]]))

    make_top(ws, number_of_rows > number_of_rows_per_page)
    left_row, right_row = write_categories(ws, order_info)
    write_header(ws, order_info)
    write_footer(ws, order_info)

    max_name_width = max([len(row['name']) * 2.62 for row in order_info['items']])

    format_xlsx(ws, max_name_width, number_of_rows < number_of_rows_per_page)

    # write_note(ws, order_info, max(left_row, right_row) + 1)

    wb.save(filename=dest_dir)


def make_top(worksheet, two_cols):

    for col in range(1, 10) if two_cols else range(1, 5):
        for row in [1, 2]:
            cell = worksheet.cell(column=col, row=row)
            cell.fill = top_color
            cell.border = box_border
    for col in [1, 6] if two_cols else [1]:
        wt_cell = worksheet.cell(column=col, row=2)
        wt_cell.value = "QTY"
        wt_cell.font = medium_font
        wt_cell.alignment = center
    for col in [2, 7] if two_cols else [2]:
        pro_cell = worksheet.cell(column=col, row=2)
        pro_cell.value = "PRODUCT"
        pro_cell.font = large_font
        pro_cell.alignment = center
    for col in [3, 8] if two_cols else [3]:
        for row_index in [1, 2]:
            weight_cell = worksheet.cell(column=col, row=row_index)
            if row_index == 2:
                weight_cell.value = "OZ"
            else:
                weight_cell.value = "WT."
            weight_cell.font = small_font
            weight_cell.alignment = center
    for col in [4, 9] if two_cols else [4]:
        for row_index in [1, 2]:
            pack_cell = worksheet.cell(column=col, row=row_index)
            if row_index == 1:
                pack_cell.value = "PACK"
                pack_cell.font = small_font
                worksheet.merge_cells(
                    start_row=1,
                    start_column=col,
                    end_row=1,
                    end_column=col + 1)
            else:
                pack_cell.value = "UPC"
                pack_cell.font = medium_font
            pack_cell.alignment = center
    for col in [5, 10] if two_cols else [5]:
        case_cell = worksheet.cell(column=col, row=2)
        case_cell.value = "CASE"
        case_cell.font = small_font
        case_cell.alignment = center


def write_categories(worksheet, order_info):
    categories = make_categories(order_info)
    rows_height = math.ceil(get_order_rows(order_info, categories)/2) + 3
    left_row = 3
    right_row = 3
    number_of_rows = len(order_info['items']) + len(set([category["type"] for category in order_info["items"]]))
    if number_of_rows < number_of_rows_per_page:
        for category in categories:
            left_row = write_category(worksheet, categories[category], left_row, "left")
    else:
        for category in categories:
            # if (left_row <= rows_height) and (len(categories[category])/2 < rows_height - left_row):
            if (left_row <= right_row):
                left_row = write_category(
                    worksheet, categories[category], left_row, "left")
            else:
                right_row = write_category(
                    worksheet, categories[category], right_row, "right")
    return (left_row, right_row)


def make_categories(order_info):
    categories = {}

    for item in order_info['items']:
        if 'type' not in item:
            continue

        if item['type'] in categories:
            categories[item['type']].append(item)
        else:
            categories[item['type']] = [item]

    return categories


def get_order_rows(order_info, order_catagories):
    return len(order_info['items']) + 2*len(order_catagories)


def write_category(worksheet, category, row_index, column):
    if column == "left":
        col = 1
    if column == "right":
        col = 6
    cat = category[0]['type']
    category_cell = worksheet.cell(column=col + 1, row=row_index)
    category_cell.value = cat
    category_cell.font = category_font
    category_cell.alignment = center
    category_cell.fill = category_color

    for cols in range(col, col + 5):
        cell = worksheet.cell(column=cols, row=row_index)
        cell.border = box_border
        cell.fill = category_color

    row_increase = 1
    for item in category:
        for col_increase in range(5):
            cell = worksheet.cell(column=col + col_increase,
                                  row=row_index + row_increase)
            cell.font = column_fonts[col_increase]

            cell.border = side_border
            if col_increase == 0:
                cell.value = item['amount']
            elif col_increase == 1:
                cell.value = item['name']
            elif col_increase == 2:
                if 'oz' in item:
                    cell.value = item['oz']
                else:
                    cell.value = 0
            elif col_increase == 3:
                cell.value = item['upc']
            elif col_increase == 4:
                if 'case' in item:
                    cell.value = item['case']
                elif 'tray' in item:
                    cell.value = item['tray']
            if col_increase != 1:
                cell.alignment = center
        row_increase += 1

    for cols in range(col, col + 5):
        cell = worksheet.cell(column=cols, row=row_index + row_increase)
        cell.border = side_with_bot_border

    return row_index + row_increase + 1


def format_date(date_iso):
    delivery_date = date(int(date_iso[0:4]), int(
        date_iso[5:7]), int(date_iso[8:10]))
    delivery_date_str = delivery_date.strftime("%a")
    if is_next_week(delivery_date):
        delivery_date_str = delivery_date.strftime("%a %m/%d")
    return delivery_date_str


def get_order_date():
    timezone = pytz.timezone("America/Anchorage")
    time_str = str(timezone.localize(datetime.now()).time())
    am_pm = "AM"
    hours = int(time_str[0:2])
    if hours > 12:
        am_pm = "PM"
        hours = hours % 12
    elif hours == 12:
        am_pm = "PM"
    minutes = time_str[3:5]
    return f'System Time: {format_date(str(date.today())).upper()} {str(hours)}:{minutes} {am_pm}\n Printed: &[Date] &[Time]'



def is_next_week(delivery_date):
    today = date.today()
    return ((delivery_date - today) +
            timedelta(days=today.weekday()) >= timedelta(days=7))


def write_header(worksheet, order_info):
    worksheet.oddHeader.left.text = get_order_date()
    worksheet.oddHeader.left.size = "18"
    worksheet.oddHeader.left.font = "Arial"
    worksheet.oddHeader.center.text = "CUSTOMER: " + \
        order_info['store']
    worksheet.oddHeader.center.size = "22"
    worksheet.oddHeader.center.font = "Arial"
    worksheet.oddHeader.right.text = "DATE ORDER TAKEN: " + str(date.today())
    worksheet.oddHeader.right.size = "18"
    worksheet.oddHeader.right.font = "Arial"


def write_footer(worksheet, order_info):
    num_rows = len(order_info['notes'].split('\n'))

    comprehension_notes = '\n'
    left_notes = comprehension_notes.join(
        [line for num, line in enumerate(order_info['notes'].split('\n')) if num <= math.ceil(num_rows / 2)-1])
    right_notes = comprehension_notes.join(
        [line for num, line in enumerate(order_info['notes'].split('\n')) if num > math.ceil(num_rows / 2)-1])

    worksheet.oddFooter.left.text = "Notes: " + left_notes
    worksheet.oddFooter.left.size = "22"
    worksheet.oddFooter.left.font = "Arial"
    worksheet.oddFooter.center.text = right_notes
    worksheet.oddFooter.center.size = "22"
    worksheet.oddFooter.center.font = "Arial"
    worksheet.oddFooter.right.text = "FOR DELIVERY ON: " + \
        format_date(order_info['date']).upper()
    worksheet.oddFooter.right.size = "22"
    worksheet.oddFooter.right.font = "Arial"



def write_note(worksheet, order_info, row_index):
    number_of_rows = len(order_info['items']) + len(set([category["type"] for category in order_info["items"]]))
    for col in range(1, 10) if number_of_rows > number_of_rows_per_page else range(1, 5):
        cell = worksheet.cell(column=col, row=row_index)
        cell.fill = category_color
    cell = worksheet.cell(column=1, row=row_index)
    cell.value = "NOTE:"
    cell.font = large_font

    cell = worksheet.cell(column=1, row=row_index + 1)
    note = list(order_info['notes'])
    index = 0
    lineLength = 0
    rows = 2
    while index < len(note):
        if note[index] == '\n':
            rows += 1
            lineLength = 0
        elif lineLength == 100:
            while note[index] != " " and note[index] != "\n":
                lineLength += 1
                index += 1
            note[index] = '\n'
            lineLength = 0
            rows += 1
        lineLength += 1
        index += 1
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    cell.alignment = wrap_alignment
    cell.value = "".join(note)
    worksheet.merge_cells(start_row=row_index + 1, start_column=1,
                          end_row=row_index + rows, end_column=10 if number_of_rows > number_of_rows_per_page else 5)
    cell.font = note_font


def format_xlsx(worksheet, max_name_width, single_column):
    col_widths = {'A': 8.08984375,
                  'B': max_name_width,
                  'C': 8.7265625,
                  'D': 19.90625,
                  'E': 10.08984375,
                  'F': 8.08984375,
                  'G': max_name_width,
                  'H': 9.1796875,
                  'I': 19.90625,
                  'J': 9.453125,
                  'K': 6.0,
                  'L': 32.81640625,
                  'M': 17.453125,
                  'N': 9.1796875}

    for column in col_widths.keys():
        worksheet.column_dimensions[column].width = col_widths[column]

    worksheet.page_setup.orientation = 'portrait'
    worksheet.page_setup.fitToPage = True
    worksheet.page_setup.fitToHeight = 1 if single_column else 2
    worksheet.page_setup.copies = 2

    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.75
    worksheet.page_margins.bottom = 0.75
    worksheet.page_margins.header = 0.3
    worksheet.page_margins.footer = 0.3
